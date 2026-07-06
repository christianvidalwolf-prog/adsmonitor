import pandas as pd
import openpyxl
import os

# Paths
sales_path = '/Users/christianvidalwolf/incremento sales/DE Sales 60 dias.xlsx'
inventory_path = '/Users/christianvidalwolf/incremento sales/Inventory DE.txt'
template_path = '/Users/christianvidalwolf/incremento sales/PriceAndQuantityDE.xlsm'

print("Loading Germany sales data...")
df_sales = pd.read_excel(sales_path)

def normalize_sku(sku, provider):
    s = str(sku).strip().upper()
    if s.startswith('AMZN.GR.'):
        s = s[len('AMZN.GR.'):]
    for suffix in ['FBA', 'FBM', 'FBS']:
        if s.endswith(suffix):
            s = s[:-len(suffix)]
            
    if provider == 'Minerales':
        if 'VC' in s:
            idx = s.find('VC')
            s = s[:idx+2]
    elif provider == 'Signes':
        if 'SG' in s:
            idx = s.find('SG')
            s = s[:idx+2]
    elif provider == 'Dcasa':
        for t in ['CLM', 'DC']:
            if t in s:
                idx = s.find(t)
                s = s[:idx+len(t)]
                break
    elif provider == 'Trediser':
        for t in ['MD', 'TD']:
            if t in s:
                idx = s.find(t)
                s = s[:idx+len(t)]
                break
    return s

# Load active inventory from Inventory DE.txt
print("Loading Germany active inventory...")
inventory = []
with open(inventory_path, 'r', encoding='utf-8-sig') as f:
    headers_line = f.readline()
    for line in f:
        parts = line.strip().split('\t')
        if len(parts) < 4:
            continue
        
        sku = parts[0].strip()
        price_str = parts[2].strip()
        qty_str = parts[3].strip()
        
        # We only keep rows with standard merchant quantity > 0
        if not qty_str or qty_str == "" or qty_str == "None":
            continue
            
        try:
            qty = int(qty_str)
            price = float(price_str)
        except ValueError:
            continue
            
        if qty <= 0:
            continue
            
        inventory.append({
            'sku': sku,
            'price': price,
            'quantity': qty
        })
print(f"Loaded {len(inventory)} active inventory rows from Inventory DE.txt.")

def generate_provider_de_file(name, filter_func, output_filename):
    print(f"\n--- Generating Germany file for {name} ---")
    
    # 1. Group sales for this provider
    df_sales_prov = df_sales.copy()
    df_sales_prov['SKU_norm'] = df_sales_prov['SKU'].apply(lambda x: normalize_sku(x, name))
    sales_map = df_sales_prov.groupby('SKU_norm')['Units'].sum().to_dict()
    print(f"Loaded {len(sales_map)} sales records for {name} in DE.")
    
    # 2. Filter inventory items
    matching_items = []
    for item in inventory:
        sku = item['sku']
        price = item['price']
        qty = item['quantity']
        
        if filter_func(sku, price):
            matching_items.append(item)
            
    print(f"Found {len(matching_items)} active items matching {name} criteria in DE.")
    
    # 3. Load template
    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    ws = wb['Vorlage']
    max_row = ws.max_row
    max_col = ws.max_column
            
    # 5. Build updated rows
    updated_rows = []
    count_0_sales = 0
    count_gt0_sales = 0
    
    for item in matching_items:
        sku = item['sku']
        std_price = item['price']
        qty = item['quantity']
        
        sku_norm = normalize_sku(sku, name)
        units_sold = sales_map.get(sku_norm, 0)
        
        # Apply B2C discount rules (identical to Spain)
        if units_sold > 0:
            multiplier = 0.93
            count_gt0_sales += 1
        else:
            multiplier = 0.75
            count_0_sales += 1
            
        sale_price = round(std_price * multiplier, 2)
        min_price = round(std_price * 0.5, 2)
        max_price = round(std_price * 2.0, 2)
        
        # Construct template row (32 columns)
        row_vals = [None] * max_col
        row_vals[0] = sku
        row_vals[1] = 'DEFAULT'
        row_vals[2] = qty
        row_vals[3] = 2
        row_vals[6] = std_price
        row_vals[8] = min_price
        row_vals[9] = max_price
        row_vals[10] = '2026-08-31'
        row_vals[11] = '2026-07-02'
        row_vals[12] = sale_price
        row_vals[13] = '2026-07-02'
        row_vals[14] = '2026-08-31'
        
        updated_rows.append(row_vals)
        
    print(f"Stats for {name}:")
    print(f"  0 sales (25% discount): {count_0_sales}")
    print(f"  >0 sales (7% discount): {count_gt0_sales}")
    print(f"  Total rows: {len(updated_rows)}")
    
    # 6. Clear original data starting from row 7
    if max_row >= 7:
        ws.delete_rows(7, max_row - 6)
        
    # 7. Write new rows
    for r_idx, row in enumerate(updated_rows, start=7):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = val
            
    # Save output file
    output_path = os.path.join('/Users/christianvidalwolf/incremento sales', output_filename)
    print(f"Saving workbook to {output_path}...")
    wb.save(output_path)
    print("Workbook saved successfully!")
    return {
        "name": name,
        "count": len(updated_rows),
        "sales_0": count_0_sales,
        "sales_gt0": count_gt0_sales,
        "filename": output_filename
    }

# Run generation for each provider
results = []

# Minerales: SKU has "VC"
results.append(generate_provider_de_file(
    'Minerales',
    lambda sku, price: 'VC' in sku.upper(),
    'PriceAndQuantity_Ofertas_Minerales_DE.xlsm'
))

# Signes: SKU has "SG" and price > 20
results.append(generate_provider_de_file(
    'Signes',
    lambda sku, price: 'SG' in sku.upper() and price > 20.0,
    'PriceAndQuantity_Ofertas_Signes_DE.xlsm'
))

# Dcasa: SKU has "CLM" or "DC" and price > 20
results.append(generate_provider_de_file(
    'Dcasa',
    lambda sku, price: ('CLM' in sku.upper() or 'DC' in sku.upper()) and price > 20.0,
    'PriceAndQuantity_Ofertas_Dcasa_DE.xlsm'
))

# Trediser: SKU has "MD" or "TD" and price > 20
results.append(generate_provider_de_file(
    'Trediser',
    lambda sku, price: ('MD' in sku.upper() or 'TD' in sku.upper()) and price > 20.0,
    'PriceAndQuantity_Ofertas_Trediser_DE.xlsm'
))

print("\n--- All Germany files generated successfully! ---")
for r in results:
    print(f"Generated {r['filename']} with {r['count']} rows ({r['sales_0']} at 25%, {r['sales_gt0']} at 7%)")
