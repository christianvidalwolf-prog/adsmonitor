import pandas as pd
import openpyxl
import os

sales_path = '/Users/christianvidalwolf/incremento sales/ventas 60 dias.xlsx'
template_path = '/Users/christianvidalwolf/incremento sales/PriceAndQuantity.xlsm'

print("Loading sales data...")
df_sales = pd.read_excel(sales_path)

def normalize_sku(sku, targets):
    s = str(sku).strip().upper()
    if s.startswith('AMZN.GR.'):
        s = s[len('AMZN.GR.'):]
    for suffix in ['FBA', 'FBM', 'FBS']:
        if s.endswith(suffix):
            s = s[:-len(suffix)]
    
    for t in targets:
        if t in s:
            idx = s.find(t)
            s = s[:idx+len(t)]
            break
    return s

def process_provider(name, targets, output_path):
    print(f"\n--- Processing {name} (keys: {targets}) ---")
    
    # Group sales
    df_sales_prov = df_sales.copy()
    df_sales_prov['SKU_norm'] = df_sales_prov['SKU'].apply(lambda x: normalize_sku(x, targets))
    sales_map = df_sales_prov.groupby('SKU_norm')['Units'].sum().to_dict()
    print(f"Loaded {len(sales_map)} normalized SKUs with sales for {name}.")
    
    # Load template
    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    ws = wb['Plantilla']
    max_row = ws.max_row
    max_col = ws.max_column
    
    # Read headers (rows 1-6) and data (rows 7+)
    headers = []
    for r in range(1, 7):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        headers.append(row_vals)
        
    data_rows = []
    for r in range(7, max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        data_rows.append(row_vals)
        
    print(f"Read {len(data_rows)} data rows from template.")
    
    # Filter and update matching rows
    updated_rows = []
    skipped_non_match = 0
    skipped_under_20 = 0
    count_0_sales = 0
    count_gt0_sales = 0
    
    for row in data_rows:
        sku = row[0]
        if sku is None:
            continue
            
        sku_str = str(sku).strip().upper()
        # Check if SKU matches targets
        is_match = any(t in sku_str for t in targets)
        if not is_match:
            skipped_non_match += 1
            continue
            
        # Get standard price
        std_price_val = row[6]
        if std_price_val is None:
            continue
            
        try:
            std_price = float(str(std_price_val).replace(',', '.'))
        except ValueError:
            continue
            
        # Check price limit
        if std_price <= 20.0:
            skipped_under_20 += 1
            continue
            
        sku_norm = normalize_sku(sku_str, targets)
        units_sold = sales_map.get(sku_norm, 0)
        
        # Apply discount
        if units_sold > 0:
            multiplier = 0.93
            count_gt0_sales += 1
        else:
            multiplier = 0.75
            count_0_sales += 1
            
        sale_price = round(std_price * multiplier, 2)
        
        # Update row values
        row[1] = 'DEFAULT'
        row[10] = '2026-08-31'
        row[11] = '2026-07-02'
        row[12] = sale_price
        row[13] = '2026-07-02'
        row[14] = '2026-08-31'
        
        updated_rows.append(row)
        
    print(f"Filtered statistics:")
    print(f"  Skipped non-{name} products: {skipped_non_match}")
    print(f"  Skipped {name} products with price <= 20 EUR: {skipped_under_20}")
    print(f"  Updated products with 0 sales (25% discount): {count_0_sales}")
    print(f"  Updated products with >0 sales (7% discount): {count_gt0_sales}")
    print(f"  Total updated products for {name}: {len(updated_rows)}")
    
    # Clear original data rows from sheet
    if max_row >= 7:
        ws.delete_rows(7, max_row - 6)
        
    # Write updated rows back
    for r_idx, row in enumerate(updated_rows, start=7):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = val
            
    print(f"Saving workbook to {output_path}...")
    wb.save(output_path)
    print("Workbook saved successfully!")

# Process Dcasa
process_provider('Dcasa', ['CLM', 'DC'], '/Users/christianvidalwolf/incremento sales/PriceAndQuantity_Ofertas_Dcasa.xlsm')

# Process Trediser
process_provider('Trediser', ['MD', 'TD'], '/Users/christianvidalwolf/incremento sales/PriceAndQuantity_Ofertas_Trediser.xlsm')
