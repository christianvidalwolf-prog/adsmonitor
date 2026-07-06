import pandas as pd
import openpyxl
import os
from datetime import datetime

# Define file paths
sales_path = '/Users/christianvidalwolf/incremento sales/ventas 60 dias.xlsx'
template_path = '/Users/christianvidalwolf/incremento sales/PriceAndQuantity.xlsm'
output_path = '/Users/christianvidalwolf/incremento sales/PriceAndQuantity_Ofertas_Minerales.xlsm'

print("Loading sales data...")
df_sales = pd.read_excel(sales_path)

# Normalize SKU function
def normalize_sku(sku):
    s = str(sku).strip().upper()
    if s.startswith('AMZN.GR.'):
        s = s[len('AMZN.GR.'):]
    for suffix in ['FBA', 'FBM', 'FBS']:
        if s.endswith(suffix):
            s = s[:-len(suffix)]
    if 'VC' in s:
        idx = s.find('VC')
        s = s[:idx+2]
    return s

df_sales['SKU_norm'] = df_sales['SKU'].apply(normalize_sku)
sales_map = df_sales.groupby('SKU_norm')['Units'].sum().to_dict()
print(f"Loaded {len(sales_map)} normalized SKUs with sales.")

print("Loading template workbook (this might take a few seconds)...")
wb = openpyxl.load_workbook(template_path, keep_vba=True)
ws = wb['Plantilla']
max_row = ws.max_row
max_col = ws.max_column
print(f"Template loaded. Sheet dimensions: {max_row} rows x {max_col} columns.")

# Read header rows (1 to 5) and example row (6)
# Note: Excel rows are 1-indexed.
headers = []
for r in range(1, 7):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
    headers.append(row_vals)

# Read data rows (7 onwards)
data_rows = []
for r in range(7, max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
    data_rows.append(row_vals)
print(f"Read {len(data_rows)} data rows from template.")

# Filter and update VC rows
updated_rows = []
skipped_non_vc = 0
count_0_sales = 0
count_gt0_sales = 0

for row in data_rows:
    sku = row[0]
    if sku is None:
        continue
    
    sku_str = str(sku).strip()
    if 'VC' not in sku_str.upper():
        skipped_non_vc += 1
        continue  # Option A: exclude non-VC products
    
    sku_norm = normalize_sku(sku_str)
    units_sold = sales_map.get(sku_norm, 0)
    
    # Get standard price
    std_price_val = row[6]  # Column G (1-indexed 7, 0-indexed 6)
    if std_price_val is None:
        print(f"Warning: Standard price is missing for SKU {sku_str}")
        continue
        
    try:
        std_price = float(str(std_price_val).replace(',', '.'))
    except ValueError:
        print(f"Warning: Could not parse standard price '{std_price_val}' for SKU {sku_str}")
        continue
        
    # Apply discount
    if units_sold > 0:
        multiplier = 0.93
        discount_label = "7%"
        count_gt0_sales += 1
    else:
        multiplier = 0.75
        discount_label = "25%"
        count_0_sales += 1
        
    sale_price = round(std_price * multiplier, 2)
    
    # Update row values
    # Col 1 is fulfillment channel code
    # Col 10 is sale end date (discounted_price...end_at)
    # Col 11 is sale start date (discounted_price...start_at)
    # Col 12 is sale price (Column M, 1-indexed 13, 0-indexed 12)
    # Col 13 is start date (Column N, 1-indexed 14, 0-indexed 13)
    # Col 14 is end date (Column O, 1-indexed 15, 0-indexed 14)
    row[1] = 'DEFAULT'
    row[10] = '2026-08-31'
    row[11] = '2026-07-02'
    row[12] = sale_price
    row[13] = '2026-07-02'
    row[14] = '2026-08-31'
    
    updated_rows.append(row)

print(f"\nFiltered statistics:")
print(f"Skipped non-VC products: {skipped_non_vc}")
print(f"Updated VC products with 0 sales (25% discount): {count_0_sales}")
print(f"Updated VC products with >0 sales (7% discount): {count_gt0_sales}")
print(f"Total updated VC products: {len(updated_rows)}")

# Clear original data rows from sheet
if max_row >= 7:
    ws.delete_rows(7, max_row - 6)

# Write updated rows back
for r_idx, row in enumerate(updated_rows, start=7):
    for c_idx, val in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.value = val

print(f"Saving workbook to {output_path}...")
wb.save(output_path)
print("Workbook saved successfully!")
